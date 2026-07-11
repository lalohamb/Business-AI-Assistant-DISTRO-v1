#!/usr/bin/env python3
"""Index Obsidian vault and system/client files into PostgreSQL + pgvector."""

import os
import csv
import email
import io
import psycopg2
from pathlib import Path
from dotenv import load_dotenv

env_path = Path(__file__).resolve().parent.parent / ".env"
load_dotenv(env_path)

BASE_PATH = os.getenv("BASE_PATH")
ACTIVE_CLIENT = os.getenv("ACTIVE_CLIENT", "my-business")
EMBEDDING_PROVIDER = os.getenv("EMBEDDING_PROVIDER", "ollama")
EMBEDDING_MODEL = os.getenv("EMBEDDING_MODEL", "nomic-embed-text")
OLLAMA_BASE_URL = os.getenv("OLLAMA_BASE_URL", "http://localhost:11434")

EXCLUDE_DIRS = {"admin", "logs", "backups", "docker", "postgres", "node_modules", ".git", "venv", ".obsidian"}
EXCLUDE_EXTENSIONS = {".key", ".pem"}
EXCLUDE_FILES = {".env"}
INDEXABLE_EXTENSIONS = {".md", ".txt", ".pdf", ".docx", ".xlsx", ".csv", ".html", ".htm", ".eml"}

INDEX_PATHS = [
    os.path.join(BASE_PATH, "system"),
    os.path.join(BASE_PATH, "clients", ACTIVE_CLIENT),
    os.path.join(BASE_PATH, "vault"),
]

DB_CONFIG = {
    "host": "localhost",
    "port": 5432,
    "user": "admin",
    "password": "strongpassword",
    "dbname": "businessassistant",
}


def get_files(paths):
    """Collect all indexable files, excluding admin/logs/backups/docker/postgres/.git."""
    files = []
    for base in paths:
        if not os.path.exists(base):
            continue
        for root, dirs, filenames in os.walk(base):
            dirs[:] = [d for d in dirs if d not in EXCLUDE_DIRS]
            for f in filenames:
                if f in EXCLUDE_FILES:
                    continue
                ext = os.path.splitext(f)[1].lower()
                if ext in EXCLUDE_EXTENSIONS:
                    continue
                if ext in INDEXABLE_EXTENSIONS:
                    files.append(os.path.join(root, f))
    return files


def extract_text(filepath):
    """Extract text content from supported file formats."""
    ext = os.path.splitext(filepath)[1].lower()

    if ext in (".md", ".txt"):
        with open(filepath, "r", errors="ignore") as f:
            return f.read().strip()

    elif ext == ".pdf":
        import fitz  # pymupdf
        doc = fitz.open(filepath)
        text = "\n".join(page.get_text() for page in doc)
        doc.close()
        return text.strip()

    elif ext == ".docx":
        from docx import Document
        doc = Document(filepath)
        return "\n".join(p.text for p in doc.paragraphs).strip()

    elif ext == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(filepath, read_only=True, data_only=True)
        lines = []
        for sheet in wb.worksheets:
            lines.append(f"## {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                lines.append(" | ".join(cells))
        wb.close()
        return "\n".join(lines).strip()

    elif ext == ".csv":
        with open(filepath, "r", errors="ignore", newline="") as f:
            reader = csv.reader(f)
            lines = [" | ".join(row) for row in reader]
        return "\n".join(lines).strip()

    elif ext in (".html", ".htm"):
        from bs4 import BeautifulSoup
        with open(filepath, "r", errors="ignore") as f:
            soup = BeautifulSoup(f.read(), "html.parser")
        return soup.get_text(separator="\n", strip=True)

    elif ext == ".eml":
        with open(filepath, "rb") as f:
            msg = email.message_from_binary_file(f)
        parts = []
        if msg["subject"]:
            parts.append(f"Subject: {msg['subject']}")
        if msg["from"]:
            parts.append(f"From: {msg['from']}")
        if msg["date"]:
            parts.append(f"Date: {msg['date']}")
        for part in msg.walk():
            if part.get_content_type() == "text/plain":
                payload = part.get_payload(decode=True)
                if payload:
                    parts.append(payload.decode(errors="ignore"))
        return "\n".join(parts).strip()

    return ""


def chunk_text(text, chunk_size=512, overlap=64):
    """Split text into overlapping chunks."""
    chunks = []
    start = 0
    while start < len(text):
        end = start + chunk_size
        chunks.append(text[start:end])
        start += chunk_size - overlap
    return [c for c in chunks if c.strip()]


def get_embedding(text):
    """Get embedding vector from configured provider."""
    if EMBEDDING_PROVIDER == "ollama":
        import requests
        resp = requests.post(
            f"{OLLAMA_BASE_URL}/api/embeddings",
            json={"model": EMBEDDING_MODEL, "prompt": text},
        )
        resp.raise_for_status()
        return resp.json()["embedding"]
    else:
        raise NotImplementedError(f"Embedding provider \"{EMBEDDING_PROVIDER}\" not yet supported.")


def index():
    """Main indexing pipeline."""
    files = get_files(INDEX_PATHS)
    print(f"Found {len(files)} files to index.")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("DELETE FROM rag_chunks WHERE client_name = %s", (ACTIVE_CLIENT,))
    cur.execute("DELETE FROM rag_documents WHERE client_name = %s", (ACTIVE_CLIENT,))

    for filepath in files:
        try:
            content = extract_text(filepath)
        except Exception as e:
            print(f"  Skipped (parse error): {filepath} — {e}")
            continue

        if not content:
            continue

        title = os.path.basename(filepath)
        rel_path = os.path.relpath(filepath, BASE_PATH)

        cur.execute(
            "INSERT INTO rag_documents (client_name, source_path, title) VALUES (%s, %s, %s) RETURNING id",
            (ACTIVE_CLIENT, rel_path, title),
        )
        doc_id = cur.fetchone()[0]

        chunks = chunk_text(content)
        for chunk in chunks:
            try:
                embedding = get_embedding(chunk)
            except Exception as e:
                print(f"  Embedding failed for chunk in {title}: {e}")
                continue

            cur.execute(
                "INSERT INTO rag_chunks (document_id, client_name, source_path, title, chunk_text, embedding) VALUES (%s, %s, %s, %s, %s, %s)",
                (doc_id, ACTIVE_CLIENT, rel_path, title, chunk, embedding),
            )

        print(f"  Indexed: {rel_path} ({len(chunks)} chunks)")

    conn.commit()
    cur.close()
    conn.close()
    print("Indexing complete.")


if __name__ == "__main__":
    index()
